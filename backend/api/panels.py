"""Panel-owner catalog — filter prospects with detected solar panels.

Exposes list + xlsx export for sales follow-up on existing panel installations.
"""

from __future__ import annotations

import asyncio
import io
import json
from datetime import UTC, datetime
from typing import Any

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel, Field
from services import geocode as geocode_mod
from services import osm_solar

from api.prospects import db

router = APIRouter()

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

CATALOG_COLUMNS: tuple[tuple[str, str], ...] = (
    ("id", "ID"),
    ("address", "Adress"),
    ("owner_name", "Ägare"),
    ("owner_age", "Ålder"),
    ("owner_phone", "Telefon"),
    ("panel_confidence", "Konfidens"),
    ("score", "Score"),
    ("annual_kwh", "kWh/år"),
    ("status", "Status"),
    ("detected_at", "Detekterad"),
    ("lat", "Lat"),
    ("lng", "Lng"),
    ("notes", "Reasoning"),
)


def _catalog_rows(min_confidence: float, limit: int) -> list[dict[str, Any]]:
    with db() as conn:
        return [
            dict(r)
            for r in conn.execute(
                """SELECT id, address, owner_name, owner_age, owner_phone,
                          panel_confidence, score, annual_kwh, status,
                          detected_at, lat, lng, notes
                   FROM prospects
                   WHERE has_panels = 1
                     AND COALESCE(panel_confidence, 0) >= ?
                   ORDER BY panel_confidence DESC NULLS LAST,
                            detected_at DESC NULLS LAST
                   LIMIT ?""",
                (min_confidence, limit),
            )
        ]


@router.get("/catalog")
async def catalog(
    min_confidence: float = Query(0.6, ge=0.0, le=1.0),
    limit: int = Query(5000, ge=1, le=50000),
) -> dict[str, Any]:
    """List buildings confirmed to have solar panels (sales-ready)."""
    rows = await asyncio.to_thread(_catalog_rows, min_confidence, limit)
    return {"count": len(rows), "min_confidence": min_confidence, "items": rows}


@router.get("/catalog.xlsx")
async def catalog_xlsx(
    min_confidence: float = Query(0.6, ge=0.0, le=1.0),
    limit: int = Query(50000, ge=1, le=100000),
) -> StreamingResponse:
    """Download panel-owner catalog as xlsx for offline sales prep."""
    rows = await asyncio.to_thread(_catalog_rows, min_confidence, limit)

    wb = Workbook()
    ws = wb.active
    ws.title = "Panel-ägare"

    headers = [label for _, label in CATALOG_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(key) for key, _ in CATALOG_COLUMNS])

    for idx, (key, _) in enumerate(CATALOG_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = (
            18 if key == "address" else 14 if key == "notes" else 12
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers_out = {
        "Content-Disposition": 'attachment; filename="panel-agare-katalog.xlsx"'
    }
    return StreamingResponse(buf, media_type=XLSX_MIME, headers=headers_out)


class OsmImportRequest(BaseModel):
    municipality: str | None = None
    lat: float | None = None
    lng: float | None = None
    radius_m: int = Field(5000, gt=0, le=50_000)
    bbox: tuple[float, float, float, float] | None = None  # (south, west, north, east)
    reverse_geocode: bool = True  # if OSM tags lack addr:*, reverse-geocode via Nominatim
    limit: int = Field(2000, gt=0, le=10_000)


@router.post("/osm-import")
async def osm_import(req: OsmImportRequest) -> dict[str, Any]:
    """Harvest OSM-tagged solar PV installations into prospects with has_panels=1.

    Registry-based: every hit is human-tagged in OpenStreetMap, so confidence=1.0.
    Bypasses Gemini entirely — free, quota-less.
    """
    if req.bbox is not None:
        sites = await osm_solar.fetch_solar_in_bbox(*req.bbox, limit=req.limit)
        area_desc = "bbox=" + ",".join(f"{x:.4f}" for x in req.bbox)
    elif req.lat is not None and req.lng is not None:
        sites = await osm_solar.fetch_solar_around(req.lat, req.lng, req.radius_m, limit=req.limit)
        area_desc = f"{req.lat:.4f},{req.lng:.4f} r={req.radius_m}m"
    elif req.municipality:
        try:
            lat, lng = await geocode_mod.geocode(req.municipality)
        except ValueError as e:
            raise HTTPException(422, f"Could not geocode municipality: {e}")
        sites = await osm_solar.fetch_solar_around(lat, lng, req.radius_m, limit=req.limit)
        area_desc = f"{req.municipality} r={req.radius_m}m"
    else:
        raise HTTPException(422, "Provide bbox, lat+lng, or municipality")

    now = datetime.now(UTC).isoformat(timespec="seconds")
    inserted = 0
    updated = 0
    skipped_no_address = 0

    with db() as conn:
        existing = {
            r[0].strip().lower(): r[1]
            for r in conn.execute("SELECT address, id FROM prospects")
        }

        for site in sites:
            address = site.address
            if not address and req.reverse_geocode:
                try:
                    address = await geocode_mod.reverse_geocode(site.lat, site.lng)
                except Exception as exc:  # noqa: BLE001 -- reverse-geocode optional, fail-soft
                    from error_logger import log_error

                    log_error(
                        "panels-reverse-geocode",
                        exc,
                        context={"lat": site.lat, "lng": site.lng, "osm_id": site.osm_id},
                    )
                    address = None
                await asyncio.sleep(1.05)  # Nominatim ≤1 req/s

            if not address:
                skipped_no_address += 1
                continue

            notes_payload = json.dumps(
                {
                    "source": "osm",
                    "osm_type": site.osm_type,
                    "osm_id": site.osm_id,
                    "capacity_kw": site.capacity_kw,
                    "method": site.method,
                }
            )

            key = address.strip().lower()
            if key in existing:
                conn.execute(
                    """UPDATE prospects
                       SET has_panels = 1,
                           panel_confidence = 1.0,
                           detected_at = ?,
                           lat = COALESCE(lat, ?),
                           lng = COALESCE(lng, ?),
                           notes = COALESCE(notes, ?),
                           updated_at = ?
                       WHERE id = ?""",
                    (now, site.lat, site.lng, notes_payload, now, existing[key]),
                )
                updated += 1
            else:
                cur = conn.execute(
                    """INSERT INTO prospects
                       (address, lat, lng, status, has_panels, panel_confidence, detected_at, notes)
                       VALUES (?, ?, ?, 'panel_owner', 1, 1.0, ?, ?)""",
                    (address, site.lat, site.lng, now, notes_payload),
                )
                existing[key] = int(cur.lastrowid)
                inserted += 1

    return {
        "area": area_desc,
        "found": len(sites),
        "inserted": inserted,
        "updated": updated,
        "skipped_no_address": skipped_no_address,
    }


def _panel_stats_sync() -> dict[str, Any]:
    with db() as conn:
        total_panels = conn.execute(
            "SELECT COUNT(*) AS n FROM prospects WHERE has_panels = 1"
        ).fetchone()["n"]
        high_conf = conn.execute(
            """SELECT COUNT(*) AS n FROM prospects
               WHERE has_panels = 1 AND COALESCE(panel_confidence, 0) >= 0.8"""
        ).fetchone()["n"]
        enriched = conn.execute(
            """SELECT COUNT(*) AS n FROM prospects
               WHERE has_panels = 1 AND owner_name IS NOT NULL"""
        ).fetchone()["n"]
    return {
        "total_panel_owners": total_panels,
        "high_confidence": high_conf,
        "contact_enriched": enriched,
    }


@router.get("/stats")
async def panel_stats() -> dict[str, Any]:
    """Quick dashboard numbers for panel-owner pipeline."""
    return await asyncio.to_thread(_panel_stats_sync)
